"""
report_generator.py
-------------------
Generates a multi-sheet Excel audit report from audit results.

Sheets produced:
  1. Summary Dashboard  – KPIs, flag counts by category, pass rate
  2. All Flagged Records – every outlier with error_type and severity
  3. HIGH Severity       – geocode + duplicate issues only
  4. MEDIUM Severity     – address, duration, status issues
  5. LOW Severity        – missing field issues
  6. Clean Records       – records that passed all checks
  7. City Breakdown      – flag counts per city (pivot)

Run via main.py or standalone after audit_engine produces results.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# ── Colour palette ────────────────────────────────────────────────────────────
NAVY   = "1B3A6B"
RED    = "C0392B"
ORANGE = "E67E22"
GREEN  = "27AE60"
YELLOW = "F39C12"
WHITE  = "FFFFFF"
LGRAY  = "F2F4F8"
MGRAY  = "D5D8DC"


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _font(bold=False, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Sheet writers ─────────────────────────────────────────────────────────────

def write_summary_sheet(ws, summary: dict):
    ws.title = "Summary Dashboard"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value          = "🚚 Transport Data Quality Audit — Summary Dashboard"
    c.font           = Font(bold=True, size=16, color=NAVY)
    c.alignment      = _center()
    c.fill           = _fill(LGRAY)
    ws.row_dimensions[1].height = 35

    # KPI blocks — row 3
    kpis = [
        ("Total Records",   summary.get("total_records", 0),    NAVY),
        ("Flagged Records", summary.get("total_flagged", 0),     RED),
        ("Clean Records",   summary.get("clean_records", 0),     GREEN),
        ("Flag Rate %",     f"{summary.get('flag_rate_pct', 0)}%", ORANGE),
    ]
    cols = [1, 2, 3, 4]
    for col, (label, val, color) in zip(cols, kpis):
        lc = ws.cell(row=3, column=col)
        vc = ws.cell(row=4, column=col)
        lc.value     = label
        lc.font      = Font(bold=True, size=10, color=WHITE)
        lc.fill      = _fill(color)
        lc.alignment = _center()
        lc.border    = _border()
        vc.value     = val
        vc.font      = Font(bold=True, size=14, color=color)
        vc.alignment = _center()
        vc.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Check breakdown table — starts row 7
    ws["A6"] = "Check Category"
    ws["B6"] = "Issues Found"
    ws["C6"] = "Severity"
    for col in ["A", "B", "C"]:
        cell = ws[f"{col}6"]
        cell.font      = Font(bold=True, color=WHITE, size=11)
        cell.fill      = _fill(NAVY)
        cell.alignment = _center()
        cell.border    = _border()

    check_rows = [
        ("Geocode Issues",    summary.get("geocode_issues",    0), "HIGH"),
        ("Duplicate IDs",     summary.get("duplicate_ids",     0), "HIGH"),
        ("Address Issues",    summary.get("address_issues",    0), "MEDIUM"),
        ("Duration Outliers", summary.get("duration_outliers", 0), "MEDIUM"),
        ("Status Issues",     summary.get("status_issues",     0), "MEDIUM"),
        ("Missing Fields",    summary.get("missing_fields",    0), "LOW"),
    ]
    severity_colors = {"HIGH": RED, "MEDIUM": ORANGE, "LOW": YELLOW}
    for r_offset, (name, count, sev) in enumerate(check_rows, start=7):
        fill_color = LGRAY if r_offset % 2 == 0 else WHITE
        ws.cell(row=r_offset, column=1, value=name).fill   = _fill(fill_color)
        ws.cell(row=r_offset, column=2, value=count).fill  = _fill(fill_color)
        sev_cell = ws.cell(row=r_offset, column=3, value=sev)
        sev_cell.fill  = _fill(severity_colors[sev])
        sev_cell.font  = Font(bold=True, color=WHITE, size=10)
        for col in [1, 2, 3]:
            ws.cell(row=r_offset, column=col).border    = _border()
            ws.cell(row=r_offset, column=col).alignment = _center()

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    # Bar chart
    chart = BarChart()
    chart.type  = "col"
    chart.title = "Issues by Check Category"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Check"
    chart.style = 10
    chart.width = 18
    chart.height = 10

    data   = Reference(ws, min_col=2, min_row=6, max_row=12)
    cats   = Reference(ws, min_col=1, min_row=7, max_row=12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E3")


def write_data_sheet(ws, df: pd.DataFrame, title: str, header_color: str):
    ws.title = title[:31]  # Excel sheet name limit
    ws.sheet_view.showGridLines = False

    if df.empty:
        ws["A1"] = "No records in this category."
        return

    cols = list(df.columns)
    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = _fill(header_color)
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(c_idx)].width = max(14, len(str(col)) + 4)

    severity_row_colors = {"HIGH": "FDECEA", "MEDIUM": "FEF9E7", "LOW": "EAF4FB"}

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        sev = getattr(row, "severity", None)
        bg  = severity_row_colors.get(sev, WHITE) if sev else (LGRAY if r_idx % 2 == 0 else WHITE)
        for c_idx, val in enumerate(row, start=1):
            cell            = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill       = _fill(bg)
            cell.alignment  = _left()
            cell.border     = _border()
            cell.font       = Font(size=9)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def write_city_pivot(ws, flagged_df: pd.DataFrame):
    ws.title = "City Breakdown"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "City-wise Flag Summary"
    ws["A1"].font      = Font(bold=True, size=13, color=NAVY)
    ws["A1"].alignment = _center()

    if flagged_df.empty or "city" not in flagged_df.columns:
        ws["A2"] = "No flagged data available."
        return

    pivot = (
        flagged_df.groupby(["city", "error_type"])
                  .size()
                  .unstack(fill_value=0)
                  .reset_index()
    )
    pivot["Total"] = pivot.iloc[:, 1:].sum(axis=1)

    headers = list(pivot.columns)
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c_idx, value=h)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = _fill(NAVY)
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(c_idx)].width = max(16, len(str(h)) + 4)

    for r_idx, row_data in enumerate(pivot.itertuples(index=False), start=3):
        bg = LGRAY if r_idx % 2 == 0 else WHITE
        for c_idx, val in enumerate(row_data, start=1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = _fill(bg)
            cell.alignment = _center()
            cell.border    = _border()
            cell.font      = Font(size=10)

    ws.freeze_panes = "A3"


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_report(results: dict, output_path: str = "reports/audit_report.xlsx"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    summary = results["summary"]
    flagged = results["flagged"]
    clean   = results["clean"]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1: Summary
        pd.DataFrame().to_excel(writer, sheet_name="Summary Dashboard", index=False)
        write_summary_sheet(writer.sheets["Summary Dashboard"], summary)

        # Sheet 2: All flagged
        if not flagged.empty:
            flagged.to_excel(writer, sheet_name="All Flagged Records", index=False)
            write_data_sheet(writer.sheets["All Flagged Records"],
                             flagged, "All Flagged Records", NAVY)

        # Sheets 3-5: By severity
        for sev, color in [("HIGH", RED), ("MEDIUM", ORANGE), ("LOW", YELLOW)]:
            sev_df = flagged[flagged["severity"] == sev] if not flagged.empty else pd.DataFrame()
            sheet_name = f"{sev} Severity"
            sev_df.to_excel(writer, sheet_name=sheet_name, index=False)
            write_data_sheet(writer.sheets[sheet_name], sev_df, sheet_name, color)

        # Sheet 6: Clean records (sample — first 2000 to keep file manageable)
        clean_sample = clean.head(2000)
        clean_sample.to_excel(writer, sheet_name="Clean Records (Sample)", index=False)
        write_data_sheet(writer.sheets["Clean Records (Sample)"],
                         clean_sample, "Clean Records (Sample)", GREEN)

        # Sheet 7: City pivot
        pd.DataFrame().to_excel(writer, sheet_name="City Breakdown", index=False)
        write_city_pivot(writer.sheets["City Breakdown"], flagged)

    # Post-process: fix summary sheet (ExcelWriter creates a blank sheet)
    wb = load_workbook(output_path)
    # Summary sheet was already written correctly via the writer reference above
    wb.save(output_path)

    print(f"[report_generator] Report saved → {output_path}")
    return output_path
